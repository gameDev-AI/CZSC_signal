#! /root/anaconda3/bin/python
# -*- coding: utf-8 -*-

import baostock as bs
import pandas as pd
import pymysql
import time
import datetime
import calendar
from datetime import timedelta
from itertools import chain

import talib 

import os
import pandas as pd
from collections import OrderedDict
from czsc import CZSC, CzscAdvancedTrader, Freq
from czsc.utils import BarGenerator
from czsc import signals
from czsc.objects import RawBar
from czsc.enum import Freq
from czsc import signals
from czsc.signals.signals import get_default_signals, get_s_three_bi, get_s_d0_bi
from czsc.signals.ta import get_s_single_k, get_s_three_k, get_s_sma, get_s_macd
from czsc.signals.bxt import get_s_like_bs, get_s_d0_bi, get_s_bi_status, get_s_di_bi, get_s_base_xt, get_s_three_bi

import new_signal
import datetime
import openpyxl
from openpyxl import load_workbook
    
import os,re,time
import pathlib
from czsc.utils.cache import home_path

os.environ['czsc_verbose'] = "1"        # 是否输出详细执行信息，0 不输出，1 输出
os.environ['czsc_min_bi_len'] = "6"     # 通过环境变量设定最小笔长度，6 对应新笔定义，7 对应老笔定义
pd.set_option('mode.chained_assignment', None)
pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', 20)

freq_map = {'1min': Freq.F1, '5min': Freq.F5, '15min': Freq.F15, '30min': Freq.F30,
            '60min': Freq.F60, 'D': Freq.D, 'W': Freq.W, 'M': Freq.M}

def float2(value, num):
    #print(round(float(value),num))
    return round(float(value),num)
    
    
def to_csv(export_signals):
    """保存到表格中"""

    # 新建文件的名称
    new_excel_file = home_path + '/30_line_stock_signals.xlsx'

    # 新建sheet的名称
    ws1_name = 'stock_signals'
    # ws2_name = 'testsheet'

    # print(new_excel_file)
    # 创建一个excel表格，默认表格内存在一个名为‘Sheet’的sheet，
    # 表格创建在缓存区，通过save(‘file_name.xlsx’)函数保存到指定目录下
    if not pathlib.Path(new_excel_file).exists():
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(new_excel_file)
    # 将默认创建的名为‘Sheet’的sheet重命名为ws1_name
    if 'Sheet' in wb.sheetnames:
        wb['Sheet'].title = ws1_name
    
    sheet1 = wb.active
    #print(sheet1)
    
    row_file = 5 # 生成5行
    col_three = 3 # 生成3列

    for row in range(len(export_signals)):
        for col in range(3):
            rw = row + 1
            cl = col + 1
            sheet1.cell(row=rw, column=cl, value=export_signals[row][col])

    # create new sheet,sheet name = ws2_name
    # wb.create_sheet(ws2_name)

    ws=wb[ws1_name]
               
    # print(new_excel_file)
    wb.save(new_excel_file)#保存 
    wb.close()    

def find_cross_point(df_stockload_list1, df_stockload_list2, c_len):
    for j in range(c_len - 1, 1, -1):
        if float2(df_stockload_list1[j],2) - float2(df_stockload_list2[j],2) < 0.002:
            return j
        else:
            continue
    return c_len
    
    
def get_test_signals(c: CZSC) -> OrderedDict:
        s = OrderedDict({"symbol": c.symbol, "dt": c.bars_raw[-1].dt, "close": c.bars_raw[-1].close})
        s.update(get_s_d0_bi(c))
        s.update(get_s_three_k(c, 1))
        s.update(get_s_di_bi(c, 1))
        s.update(get_s_macd(c, 1))
        return s
        


# 定义一些需要观察的信号，可以是多级别同时计算
def get_simple_signals(cat: CzscAdvancedTrader) -> OrderedDict:
    s = OrderedDict({"symbol": cat.symbol, "dt": cat.end_dt, "close": cat.latest_price})
    for _, c in cat.kas.items():
        if c.freq == Freq.D:
            # s.update(new_signal.get_s_base_xt(c, di=1))
            # s.update(new_signal.get_s_three_bi(c, di=1))
            # s.update(new_signal.get_s_like_bs(c, di=1))
            
            # s.update(new_signal.get_s_di_bi(c, di=1))
            # s.update(new_signal.get_s_bi_status(c))
            
            # s.update(signals.ta.get_s_single_k(c, di=1))
            # s.update(new_signal.get_s_d0_bi(c))
            # s.update(new_signal.get_s_three_k(c, di=1))
            
            #s.update(new_signal.get_s_pos(c))
            s.update(new_signal.get_bs_pos(c))
   
            # s.update(signals.ta.get_s_macd(c, di=1))
            # s.update(signals.ta.get_s_sma(c, di=1, t_seq=(5, 20, 60)))
            
            # s.update(signals.vol.get_s_vol_single_sma(c, di=1, t_seq=(10, 20)))
            # s.update(signals.vol.get_s_vol_double_sma(c, di=1, t1=5, t2=20))
            # s.update(signals.vol.get_s_amount_n(c, di=1, n=10, total_amount=10))
            
            # s.update(signals.other.get_s_zdt(c, di=1))
            
            ## s.update(signals.other.get_s_op_time_span(c, op='开多', time_span=('13:00', '14:50')))
            ## s.update(signals.other.get_s_op_time_span(c, op='平多', time_span=('09:35', '14:50')))
            ## s.update(signals.other.get_s_raw_bar_end(c, k1='60分钟'))
            
            ## s.update(signals.pos.get_s_long01(cat, th=100))
            ## s.update(signals.pos.get_s_long02(cat, th=100))
            ## s.update(signals.pos.get_s_long05(cat, span='月', th=500))

            ## if cat.long_pos:
                ## s.update(signals.cat.get_s_position(cat, cat.long_pos))
            ## if cat.short_pos:
                ## s.update(signals.cat.get_s_position(cat, cat.short_pos))
                
        # if c.freq == Freq.W:
            # s.update(new_signal.get_s_three_bi(c, di=1))
            # s.update(signals.ta.get_s_macd(c, di=1))
            # s.update(signals.ta.get_s_sma(c, di=1, t_seq=(5, 20, 60)))
    return s


def simple_strategy(symbol):
    return {"symbol": symbol, "get_signals": get_simple_signals}




def download_day_data_by_code(code, start_date, end_date):
    # 获取指定股票数据
    data_df = pd.DataFrame()
    #print(stockcodes.shape[0])
    if True:
    # for code in stockcodes["code"]:
        print("Downloading day:" + code)
        # k_rs = bs.query_history_k_data_plus("sz.003037", "date,time,code,open,high,low,close,volume,amount,adjustflag", start_date, end_date, frequency="30", adjustflag="3")
        k_rs = bs.query_history_k_data_plus(code, "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST", start_date, end_date, frequency="d", adjustflag="2")
        ##sh.600006
        # k_rs = bs.query_history_k_data_plus("sh.600006", "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST", start_date, end_date, frequency="d", adjustflag="2")
        df_stockload = k_rs.get_data()
        c_len = df_stockload.shape[0]
        freq = 'D'
        bars = []
        if c_len > 0:
            # print(df_stockload['close'][c_len - 1])
            for j in range(c_len):
                #print(df_stockload['amount'][j])
                try:
                    bars.append(RawBar(symbol=df_stockload['code'][j], dt=pd.to_datetime(df_stockload['date'][j]), id=j, freq=freq_map[freq],
                                   open=round(float(df_stockload['open'][j]), 2),
                                   close=round(float(df_stockload['close'][j]), 2),
                                   high=round(float(df_stockload['high'][j]), 2),
                                   low=round(float(df_stockload['low'][j]), 2),
                                   vol=round(float(df_stockload['volume'][j]), 2)))
                except Exception as err:
                     print(err)
                     bars = []
                    #print(str(resu[0]),str(resu[1]))
                     return bars
            #print(c_len)
            return bars
        else:
            return bars
            
    
def filter_data_by_code(stockcodes, start_date, end_date):
    # 获取指定股票数据
    data_df = pd.DataFrame()
    #print(stockcodes.shape[0])
    # if True:
    export_signals = [] 
    for code in stockcodes["code"]:
        if ("sh.0" not in code) and ("bj.8" not in code)  and ("bj.4" not in code) and ("sh.688" not in code) and ("sz.399" not in code):
            #print("Downloading day:" + code)
            # k_rs = bs.query_history_k_data_plus("sz.003037", "date,time,code,open,high,low,close,volume,amount,adjustflag", start_date, end_date, frequency="30", adjustflag="3")
            k_rs = bs.query_history_k_data_plus(code, "date,time,code,open,high,low,close,volume,amount,adjustflag", start_date, end_date, frequency="30", adjustflag="3")
            df_stockload = k_rs.get_data()
            c_len = df_stockload.shape[0]
            if c_len > 0:
                # print("Downloading day:" + code)
                # macd_dif, macd_dea, macd_bar = talib.MACD(df_stockload['close'].values, fastperiod=12, slowperiod=26, signalperiod=9)
                
                df_stockload['Ma27'] = df_stockload.close.rolling(window=27).mean()#pd.rolling_mean(df_stockload.close,window=30)
                df_stockload['Ma60'] = df_stockload.close.rolling(window=60).mean()#pd.rolling_mean(df_stockload.close,window=30)
                df_stockload['Ma108'] = df_stockload.close.rolling(window=108).mean()#pd.rolling_mean(df_stockload.close,window=30)
                df_stockload['Ma216'] = df_stockload.close.rolling(window=216).mean()#pd.rolling_mean(df_stockload.close,window=30)
                df_stockload['Ma2000'] = df_stockload.close.rolling(window=2000).mean()#pd.rolling_mean(df_stockload.close,window=30)
                
                
                close = float2(df_stockload['close'][c_len - 1],2)
                line27 = float2(df_stockload['Ma27'][c_len - 1],2)
                line60 = float2(df_stockload['Ma60'][c_len - 1],2)
                line108 = float2(df_stockload['Ma108'][c_len - 1],2)
                line216 = float2(df_stockload['Ma216'][c_len - 1],2)
                line2000 = float2(df_stockload['Ma2000'][c_len - 1],2)
                
                
                cross_point1 = find_cross_point(df_stockload['Ma27'], df_stockload['Ma60'], c_len)
                cross_point2 = find_cross_point(df_stockload['Ma27'], df_stockload['Ma108'], c_len)
                cross_point3 = find_cross_point(df_stockload['Ma27'], df_stockload['Ma216'], c_len)
                
                cross_point4 = find_cross_point(df_stockload['Ma60'], df_stockload['Ma108'], c_len)
                cross_point5 = find_cross_point(df_stockload['Ma60'], df_stockload['Ma216'], c_len)
                
                cross_point6 = find_cross_point(df_stockload['Ma108'], df_stockload['Ma216'], c_len)
                
                min_cross = cross_point3 > cross_point2 and cross_point2 > cross_point1
                med_cross = cross_point5 > cross_point4
                #max_cross = not (cross_point6 == c_len) and cross_point6 > cross_point3
                max_cross = True
                
                lineup = line27 > line60 and line27 > line108 and line27 > line216#3大于7，大于13 大于 27日均线
                
                priceup = close > line108 and close > line2000#收盘价过13日和250日均线
                
                cross_flag = not (cross_point1 == c_len) and min_cross and med_cross and max_cross
                
                if lineup and priceup and cross_flag:
                    print(df_stockload['code'][cross_point6])
                    code = df_stockload['code'][cross_point6]
                    bars = download_day_data_by_code(code, start_dt, end_dt)
                    if bars == []:
                        continue
                    else:
                        c = CZSC(bars, get_signals=get_test_signals)
                        
                        bg = BarGenerator(base_freq='日线', freqs=['周线'], max_count=5000)
                        for bar in bars:
                            bg.update(bar)
                        bars_w = bg.bars['周线']
                        
                        cat = CzscAdvancedTrader(bg, simple_strategy)
                        signals = {k: v for k, v in cat.s.items() if len(k.split("_")) == 3}
                        for k in signals.keys():
                            print(cat.symbol + " " + k + " " + signals[k])
                            export_signals.append([cat.symbol, k, signals[k]])
                        
                    #cat.open_in_browser()
        else:
            continue
    to_csv(export_signals)

    
def get_all_stock(date):
    stock_rs = bs.query_all_stock(date)
    stock_df = stock_rs.get_data()
    #print(stock_df)
    return stock_df
    

                    
def get_last_trade_day(start_dt, end_dt):
    rs = bs.query_trade_dates(start_date=start_dt, end_date=end_dt)
    print('query_trade_dates respond error_code:'+rs.error_code)
    print('query_trade_dates respond  error_msg:'+rs.error_msg)

    #### 打印结果集 ####
    data_list = []


    while (rs.error_code == '0') & rs.next():
        # 获取一条记录，将记录合并在一起
        data = rs.get_row_data()
        #print(data[1])
        if data[1] == '1':
            data_list.append(data)
        
    #result = pd.DataFrame(data_list, columns=rs.fields)

    #### 结果集输出到csv文件 ####   
    #result.to_csv("D:\\trade_datas.csv", encoding="gbk", index=False)
    #print(data_list)
    print(data_list[len(data_list) - 1][0])
    return data_list[len(data_list) - 1][0]
    
    
def get_delta_day():
        start_time = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '8:00', '%Y-%m-%d%H:%M')
        end_time = datetime.datetime.strptime(str(datetime.datetime.now().date()) + '18:00', '%Y-%m-%d%H:%M')
        # 结束时间
        now_time = datetime.datetime.now()
        # 判断当前时间是否在范围时间内
        if start_time < now_time < end_time:
            return 1
        else:
            return 1
            
if __name__ == '__main__':
    if True:       
        time_temp = datetime.datetime.now() - datetime.timedelta(days=2000)
        start_dt = time_temp.strftime('%Y-%m-%d')
        
        day = get_delta_day()
        
        time_temp = datetime.datetime.now() - datetime.timedelta(days=day)
        end_dt = time_temp.strftime('%Y-%m-%d')
        
        
        bs.login()
        
        lasttradeday = get_last_trade_day(start_dt, end_dt)
        weekday = datetime.datetime.now().isoweekday()
        
        
        # 获取全部股票的信息
        all_stock_codes = get_all_stock(lasttradeday)
        # print(all_stock_codes)
        
        total = all_stock_codes.shape[0]
        # 获取全部股票的日信息
        if total != 0:
            print("过滤出符合条件的个股")
            filter_data_by_code(all_stock_codes, start_dt, end_dt)
            
        print("handle finish")
        

