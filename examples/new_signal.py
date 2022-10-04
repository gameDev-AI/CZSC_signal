# -*- coding: utf-8 -*-
"""
author: zengbin93
email: zeng_bin8888@163.com
create_dt: 2021/11/21 17:48
describe: 笔相关信号的计算
"""
from typing import List, Union
from collections import OrderedDict

from czsc import CZSC, CzscAdvancedTrader, Freq
from czsc import analyze
from czsc.objects import Direction, BI, FakeBI, Signal
from czsc.enum import Freq
from czsc.utils.ta import RSQ

import numpy as np
from collections import OrderedDict

try:
    from czsc.utils.ta1 import MACD, SMA
except:
    from czsc.utils.ta import MACD, SMA
    
from czsc.signals.utils import is_bis_down, is_bis_up, get_zs_seq
from czsc.utils.cache import home_path

from datetime import datetime
import openpyxl
from openpyxl import load_workbook
    
import os,re,time
import pathlib

##to do list
#细化信号
#多周期关联

def find_cross_point(df_stockload_list1, df_stockload_list2, c_len):
    for j in range(c_len - 1, 1, -1):
        if float2(df_stockload_list1[j],2) - float2(df_stockload_list2[j],2) < 0.002:
            return j
        else:
            continue
    return c_len
    

def float2(value, num):
    #print(round(float(value),num))
    return round(float(value),num)

    
def check_sub_level(c: analyze.CZSC):
    if c.bi_list and len(c.bars_ubi) > 3 :
        # 表里关系的定义参考：http://blog.sina.com.cn/s/blog_486e105c01007wc1.html
        min_ubi = min([x.low for x in c.bars_ubi[3:]])
        max_ubi = max([x.high for x in c.bars_ubi[3:]])
        last_bi = c.bi_list[-1]
        if last_bi.direction == Direction.Down:
            if min_ubi > last_bi.low:##底分完成
                bis = c.finished_bis
                if not bis:
                    return True
                last_bi = bis[-1]
                if len(last_bi.bars) >= 9:# 长度达标
                    if len(last_bi.fxs) >= 6:# 小级别达标
                        return True
    return True

def get_bs_pos(c: analyze.CZSC) -> OrderedDict:
    """获取当前买卖点信号
    信号格式：周期-结构-买卖点-
    周期：30分钟，日线，周线，月线
    结构：结构
    买卖点：

    :param c: CZSC 对象
    :return: 信号字典
    """
    
    freq: Freq = c.freq
    s = OrderedDict()

    k1 = str(freq.value)
    k2 = "结构"
    k3 = "买卖点"
    v1 = "无次级别结构"
    v2 = "任意"
    v3 = "任意"
    v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
    s[v.key] = v.value
    
    if c.bi_list and len(c.bars_ubi) > 3 :
        # 表里关系的定义参考：http://blog.sina.com.cn/s/blog_486e105c01007wc1.html
        min_ubi = min([x.low for x in c.bars_ubi[3:]])
        max_ubi = max([x.high for x in c.bars_ubi[3:]])
        last_bi = c.bi_list[-1]
        if last_bi.direction == Direction.Down:
            if min_ubi > last_bi.low:##底分完成
                bis = c.finished_bis
                if not bis:
                    return s
                last_bi = bis[-1]
                if len(last_bi.bars) >= 9:# 长度达标
                    if len(last_bi.fxs) >= 6:# 小级别达标
                        bi_list = c.bi_list
                        zs_seq = get_zs_seq(bi_list)
                        zs_len = len(zs_seq)
                        if zs_len < 1:#没有中枢
                            return s
                        elif zs_len == 1:#1个中枢
                            zs = zs_seq[-1]
                            last_zs_bi_list = zs.bis
                            bi_len = len(last_zs_bi_list)
                            bi = last_zs_bi_list[-1]
                            low = min([x.low for x in last_zs_bi_list])
                            high = max([x.high for x in last_zs_bi_list])
                                
                            if bi_len % 2 == 0:#偶数笔
                                if bi_len == 2:
                                    bi1, bi2 = last_zs_bi_list
                                    if bi2.low < bi1.low and bi2.direction == Direction.Down:
                                        v1 = "背弛1买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    elif bi2.low > bi1.low and bi2.direction == Direction.Down :
                                        v1 = "标准2买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                if bi_len == 4:
                                    bi3 = last_zs_bi_list[-3]
                                    if bi.low > low and bi.low > bi3.low:
                                        v1 = "标准起爆点"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                if bi.low == low:
                                    v1 = "标准1买"
                                    v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                    s[v.key] = v.value
                                    return s
                                if bi_len >= 6:
                                    bi1 = last_zs_bi_list[-2]
                                    if bi1.low == low and bi1.high > zs.zd and bi.low > bi1.low:
                                        v1 = "强2买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                else:
                                    return s
                            else:#奇数笔
                                if bi_len > 2:
                                    bi3 = last_zs_bi_list[-3]
                                    if bi.low > low and bi.low > bi3.low:
                                        v1 = "标准起爆点"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                else:
                                    return s
                        else:#多个中枢
                            zs = zs_seq[-1]
                            zs2 = zs_seq[-2]
                            last_zs_bi_list = zs.bis
                            bi_len = len(last_zs_bi_list)
                            if bi_len == 1:
                                bi = last_zs_bi_list[-1]
                                if bi.low > zs2.zg:
                                    v1 = "二三合买"
                                    v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                    s[v.key] = v.value
                                    return s
                                else:
                                    return s
                            if bi_len % 2 == 0:#偶数笔
                                low = min([x.low for x in last_zs_bi_list])
                                high = max([x.high for x in last_zs_bi_list])
                                bi = last_zs_bi_list[-1]
                                if bi_len == 2:
                                    bi2 = last_zs_bi_list[-2]
                                    if bi.low > bi2.low:
                                        v1 = "弱二买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        bi1, bi2 = last_zs_bi_list[-2:]
                                        if bi2.low < bi1.low and bi2.direction == Direction.Down:
                                            v1 = "背弛1买"
                                            v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                            s[v.key] = v.value
                                            return s
                                        elif bi2.low > bi1.low and bi2.direction == Direction.Down :
                                            v1 = "标准2买"
                                            v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                            s[v.key] = v.value
                                            return s
                                        else:
                                            return s
                                if bi_len == 4:
                                    bi3 = last_zs_bi_list[-3]
                                    bi4 = last_zs_bi_list[-4]
                                    if bi.low > low and bi.low > bi3.low and bi4.low == low:
                                        v1 = "标准起爆点"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                if bi_len >= 6:
                                    bi1 = last_zs_bi_list[-2]
                                    if bi1.low == low and bi1.high > zs.zd and bi.low > bi1.low:
                                        v1 = "强2买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                if bi.low == low:
                                    v1 = "标准1买"
                                    v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                    s[v.key] = v.value
                                    return s
                                else:
                                    return s
                            else:#奇数笔
                                low = min([x.low for x in last_zs_bi_list])
                                if bi_len >= 4:
                                    bi1 = last_zs_bi_list[-2]
                                    bi = last_zs_bi_list[-1]
                                    bi4 = last_zs_bi_list[-4]
                                    if bi.low == low:
                                        v1 = "标准1买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    elif bi1.low == low and bi1.high > zs.zd and bi.low > bi1.low:
                                        v1 = "强2买"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    elif bi4.low == low and bi4.high > zs.zd and bi.low > bi4.low:
                                        v1 = "标准起爆点"
                                        v = Signal(k1=str(freq.value), k2=k2, k3=k3, v1=v1, v2 = v2, v3 = v3)
                                        s[v.key] = v.value
                                        return s
                                    else:
                                        return s
                                else:
                                    return s
                    else:# 小级别不达标
                        return s
                else:
                    return s
            else:
                return s
        else:
            return s
    else:
        return s
        
    

def to_csv(export_signals):
    """保存到表格中"""

    # 新建文件的名称
    new_excel_file = home_path + '/stock_signals.xlsx'

    # 新建sheet的名称
    ws1_name = 'stock_signals'
    # ws2_name = 'testsheet'

    # print(new_excel_file)
    # 创建一个excel表格，默认表格内存在一个名为‘Sheet’的sheet，
    # 表格创建在缓存区，通过save(‘file_name.xlsx’)函数保存到指定目录下
    if not pathlib.Path(new_excel_file).exists():
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(new_excel_file)
    # 将默认创建的名为‘Sheet’的sheet重命名为ws1_name
    if 'Sheet' in wb.sheetnames:
        wb['Sheet'].title = ws1_name
    
    sheet1 = wb.active
    #print(sheet1)
    
    row_file = 5 # 生成5行
    col_three = 3 # 生成3列

    for row in range(len(export_signals)):
        for col in range(3):
            rw = row + 1
            cl = col + 1
            sheet1.cell(row=rw, column=cl, value=export_signals[row][col])

    # create new sheet,sheet name = ws2_name
    # wb.create_sheet(ws2_name)

    ws=wb[ws1_name]
               
    # print(new_excel_file)
    wb.save(new_excel_file)#保存 
    wb.close()
    
    
def get_k2_status(c: analyze.CZSC) :
    bi_list = c.bi_list
    zs_seq = get_zs_seq(bi_list)
    zs_len = len(zs_seq)
    k2 = "中枢个数-中枢位置"
    if zs_len < 1:
        k2 = "无结构"
    elif zs_len == 1:
        k2 = "单中枢-构建中"
    else:
        zs_list = zs_seq[-2:]
        zs1, zs2 = zs_list
        last_zs_bi_list = zs2.bis
        last_zs_bi_len = len(last_zs_bi_list)
        if last_zs_bi_len < 2:
            last_zs_bi = last_zs_bi_list[-1]
            if last_zs_bi.direction == Direction.Up and last_zs_bi.high < zs1.zd:
                k2 = "多中枢-下方"
            elif last_zs_bi.direction == Direction.Down and last_zs_bi.low > zs1.zg:
                k2 = "多中枢-上方"
        else:
            if zs2.zg > zs1.zg:
                k2 = "多中枢-上方"
            elif zs2.zg < zs1.zg:
                k2 = "多中枢-下方"
            
    return k2
    
def get_k3_status(c: analyze.CZSC) :
    bi_list = c.bi_list
    zs_seq = get_zs_seq(bi_list)
    zs_len = len(zs_seq)
    #print(zs_len)
    k3 = "无方向-平衡"
    if zs_len < 1:
        k3 = "无方向"
    elif zs_len >= 1:
        zs = zs_seq[-1]
        last_zs_bi_list = zs.bis
        bi_len = len(last_zs_bi_list)
        last_zs_bi = last_zs_bi_list[-1]
        zs_zz = zs.zz
        bi_high = last_zs_bi.high
        bi_low = last_zs_bi.low
        direction = last_zs_bi.direction
        
        if bi_len == 1:
            k3 = "回拉-平衡"
        elif bi_len == 2:
            if bi_len == 2 and bi_high < zs_zz and direction == Direction.Up:
                k3 = "回试-弱"
            elif bi_len == 2 and bi_low > zs_zz and direction == Direction.Down:
                k3 = "回试-强"
            else:
                k3 = "回试-平衡"
        elif bi_len == 3:
            if bi_len == 3 and bi_high < zs_zz and direction == Direction.Up:
                k3 = "2次回拉-弱"
            elif bi_len == 3 and bi_low > zs_zz and direction == Direction.Down:
                k3 = "2次回拉-强" 
            else:
                k3 = "2次回拉-平衡"
        elif bi_len == 4:
            if bi_len == 4 and bi_high < zs_zz and direction == Direction.Up:
                k3 = "2次回试-弱"
            elif bi_len == 4 and bi_low > zs_zz and direction == Direction.Down:
                k3 = "2次回试-强"
            else:
                k3 = "2次回试-平衡"
                
        elif bi_len % 2 == 1:
            if bi_high < zs_zz and direction == Direction.Up:
                k3 = "多次回拉-弱"
            elif bi_low > zs_zz and direction == Direction.Down:
                k3 = "多次回拉-强" 
            else:
                k3 = "多次回拉-平衡"
        else:
            if bi_high < zs_zz and direction == Direction.Up:
                k3 = "多次回试-弱"
            elif  bi_low > zs_zz and direction == Direction.Down:
                k3 = "多次回试-强" 
            else:
                k3 = "多次回试-平衡" 
                
    return k3
    
#倒数1笔的长度和方向
def get_v1_status(c: analyze.CZSC) :
    v1 = "倒数1笔的长度-方向"
    if c.bi_list:
        # 表里关系的定义参考：http://blog.sina.com.cn/s/blog_486e105c01007wc1.html
        min_ubi = min([x.low for x in c.bars_ubi])
        max_ubi = max([x.high for x in c.bars_ubi])

        last_bi = c.bi_list[-1]
        v = None
        
        # 长度
        if len(last_bi.bars) >= 15:
            vv1="15K以上"
        elif 15 > len(c.bars_ubi) > 9:
            vv1="9-15K"
        else:
            vv1="9K以下"
        
        if last_bi.direction == Direction.Down:
            if min_ubi < last_bi.low:
                v1 = vv1 + "-向下延伸"
            else:
                v1 = vv1 + "-底分完成"
        if last_bi.direction == Direction.Up:
            if max_ubi > last_bi.high:
                v1 = vv1 + "-向上延伸"
            else:
                v1 = vv1 + "-顶分完成"
                
    return v1
    
def get_v2_status(c: analyze.CZSC) :
    #倒数未确认笔的长度和方向
    v2 = "未确认笔方向-未确认笔长度"
    bis = c.finished_bis
    if bis:
        # 倒0笔长度
        bars_ubi = [x for x in c.bars_raw[-20:] if x.dt >= bis[-1].fx_b.elements[0].dt]
        if len(bars_ubi) >= 9:
            vv2="9K以上"
        elif 9 > len(bars_ubi) > 5:
            vv2="5-9K"
        else:
            vv2="5K以下"
            
        # 倒0笔方向
        last_bi = bis[-1]
        if last_bi.direction == Direction.Down:
            v2= vv2 + "-向上"
        elif last_bi.direction == Direction.Up:
            v2= vv2 + "-向下"
        else:
            raise ValueError
                
    return v2
    
def get_v3_status(c: analyze.CZSC):
    bis = c.finished_bis
    v3 = "非买卖点"

    freq: Freq = c.freq
    if not bis:
        return v3

    if len(bis) < 3:
        return v3
    
    bi1, bi2, bi3 = bis[-3:]
    
    if not (bi1.direction == bi3.direction):
        print(f"1,3 的 direction 不一致，无法识别三笔形态，{bi3}")
        return v3

    assert bi3.direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if bi3.direction == Direction.Down:
        # 向下不重合
        if bi3.low > bi1.high:
            return '二三合买'

        # 向下奔走型
        if bi2.low < bi3.low < bi1.high < bi2.high:
            return '起爆点'

        # 向下收敛
        if bi1.high > bi3.high and bi1.low < bi3.low:
            return '二买'

        if bi1.high < bi3.high and bi1.low > bi3.low:
            return '转二卖'

        if bi3.low < bi1.low and bi3.high < bi1.high:
            if bi3.power < bi1.power:
                return '一买'
            else:
                return '转三卖'

    if bi3.direction == Direction.Up:
        if bi3.high < bi1.low:
            return '三卖'

        if bi2.low < bi1.low < bi3.high < bi2.high:
            return '二卖'

        if bi1.high > bi3.high and bi1.low < bi3.low:
            return '二卖'

        if bi1.high < bi3.high and bi1.low > bi3.low:
            return '转二买'

        if bi3.low > bi1.low and bi3.high > bi1.high:
            if bi3.power < bi1.power:
                return '卖转二买'
            else:
                return '卖转起爆点'
                
    
def get_s_pos(c: analyze.CZSC) -> OrderedDict:
    """获取当前处于中枢构建中还是移动中
    信号格式：日线_多中枢-上方-
    周期：30分钟，日线，周线，月线
    中枢信息：中枢状态--多中枢/单中枢&上移/下移/构建中
    周期_中枢状态&中枢位置_倒数1笔的长度和方向_倒数未确认笔的长度和方向

    :param c: CZSC 对象
    :param di: 最近一根K线为倒数第i根
    :return: 信号字典
    """
    
    freq: Freq = c.freq
    k1 = str(freq.value)
    k2 = "中枢个数-中枢位置"
    k3 = "确认笔数量-确认笔强弱"
    v1 = "倒数1笔的长度-方向"
    v2 = "未确认笔方向-未确认笔长度"
    v3 = ""
    
    k2 = get_k2_status(c)
    k3 = get_k3_status(c)
    
    v1 = get_v3_status(c)
    
    v2 = get_v1_status(c)
    
    v3 = get_v2_status(c)
            
    s = OrderedDict()
    v = Signal(k1=k1, k2=k2, k3=k3, v1=v1, v2=v2, v3=v3)
    s[v.key] = v.value

    return s
    

def get_s_single_k(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """获取倒数第i根K线的单K信号"""
    if c.freq not in [Freq.D, Freq.W]:
        return OrderedDict()

    if len(c.bars_raw) < di:
        return OrderedDict()

    s = OrderedDict()
    freq: Freq = c.freq
    k1 = str(freq.value)
    default_signals = [
        Signal(k1=k1, k2=f"倒{di}K", k3="状态", v1="其他", v2='其他', v3='其他'),
    ]
    for signal in default_signals:
        s[signal.key] = signal.value

    k = c.bars_raw[-di]
    if k.close > k.open:
        v = Signal(k1=k1, k2=f"倒{di}K", k3="状态", v1="上涨")
    else:
        v = Signal(k1=k1, k2=f"倒{di}K", k3="状态", v1="下跌")
    s[v.key] = v.value
    return s


def get_s_three_k(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """倒数第i根K线的三K信号

    :param c: CZSC 对象
    :param di: 最近一根K线为倒数第i根
    :return: 信号字典
    """
    assert di >= 1
    freq: Freq = c.freq
    k1 = str(freq.value)
    k2 = f"倒{di}K"

    s = OrderedDict()
    v = Signal(k1=k1, k2=k2, k3="三K形态", v1="其他", v2='其他', v3='其他')
    s[v.key] = v.value

    if len(c.bars_ubi) < 3 + di:
        return s

    if di == 1:
        tri = c.bars_ubi[-3:]
    else:
        tri = c.bars_ubi[-3 - di + 1:-di + 1]

    if tri[0].high > tri[1].high < tri[2].high:
        v = Signal(k1=k1, k2=k2, k3="三K形态", v1="底分型")
    elif tri[0].high < tri[1].high < tri[2].high:
        v = Signal(k1=k1, k2=k2, k3="三K形态", v1="向上走")
    elif tri[0].high < tri[1].high > tri[2].high:
        v = Signal(k1=k1, k2=k2, k3="三K形态", v1="顶分型")
    elif tri[0].high > tri[1].high > tri[2].high:
        v = Signal(k1=k1, k2=k2, k3="三K形态", v1="向下走")
    else:
        v = None

    if v and "其他" not in v.value:
        s[v.key] = v.value

    return s
	

def check_three_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别由远及近的三笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的三笔形态
    :param di: 最近一笔为倒数第i笔
    :return:
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='次级别无结构', v2='其他', v3='其他')

    if len(bis) != 3:
        return v

    bi1, bi2, bi3 = bis
    if not (bi1.direction == bi3.direction):
        print(f"1,3 的 direction 不一致，无法识别三笔形态，{bi3}")
        return v

    assert bi3.direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if bi3.direction == Direction.Down:
        # 向下不重合
        if bi3.low > bi1.high:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下不重合')

        # 向下奔走型
        if bi2.low < bi3.low < bi1.high < bi2.high:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下奔走型')

        # 向下收敛
        if bi1.high > bi3.high and bi1.low < bi3.low:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下收敛')

        if bi1.high < bi3.high and bi1.low > bi3.low:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下扩张')

        if bi3.low < bi1.low and bi3.high < bi1.high:
            if bi3.power < bi1.power:
                return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下盘背')
            else:
                return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向下无背')

    if bi3.direction == Direction.Up:
        if bi3.high < bi1.low:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上不重合')

        if bi2.low < bi1.low < bi3.high < bi2.high:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上奔走型')

        if bi1.high > bi3.high and bi1.low < bi3.low:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上收敛')

        if bi1.high < bi3.high and bi1.low > bi3.low:
            return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上扩张')

        if bi3.low > bi1.low and bi3.high > bi1.high:
            if bi3.power < bi1.power:
                return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上盘背')

            else:
                return Signal(k1=freq.value, k2=di_name, k3='三笔形态', v1='向上无背')
    return v


def check_five_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别五笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的五笔
    :param di: 最近一笔为倒数第i笔
    :return:
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='次级别无结构', v2='其他', v3='其他')

    if len(bis) != 5:
        return v

    bi1, bi2, bi3, bi4, bi5 = bis
    if not (bi1.direction == bi3.direction == bi5.direction):
        print(f"1,3,5 的 direction 不一致，无法识别五段形态；{bi1}{bi3}{bi5}")
        return v

    direction = bi1.direction
    max_high = max([x.high for x in bis])
    min_low = min([x.low for x in bis])
    assert direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if direction == Direction.Down:
        # aAb式底背驰
        if min(bi2.high, bi4.high) > max(bi2.low, bi4.low) and max_high == bi1.high and bi5.power < bi1.power:
            if (min_low == bi3.low and bi5.low < bi1.low) or (min_low == bi5.low):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

        # 类趋势底背驰
        if max_high == bi1.high and min_low == bi5.low and bi4.high < bi2.low and bi5.power < max(bi3.power, bi1.power):
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

        # 上颈线突破
        if (min_low == bi1.low and bi5.high > min(bi1.high, bi2.high) > bi5.low > bi1.low) \
                or (min_low == bi3.low and bi5.high > bi3.high > bi5.low > bi3.low):
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='上颈线突破')

        # 五笔三买，要求bi5.high是最高点
        if max_high == bi5.high > bi5.low > max(bi1.high, bi3.high) \
                > min(bi1.high, bi3.high) > max(bi1.low, bi3.low) > min_low:
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='类三买')

    if direction == Direction.Up:
        """aAb式类一卖"""
        if min(bi2.high, bi4.high) > max(bi2.low, bi4.low) and min_low == bi1.low and bi5.power < bi1.power:
            if (max_high == bi3.high and bi5.high > bi1.high) or (max_high == bi5.high):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

        """类趋势类一卖"""
        if min_low == bi1.low and max_high == bi5.high and bi5.power < max(bi1.power, bi3.power) and bi4.low > bi2.high:
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

        """下颈线突破"""
        if (max_high == bi1.high and bi5.low < max(bi1.low, bi2.low) < bi5.high < max_high) \
                or (max_high == bi3.high and bi5.low < bi3.low < bi5.high < max_high):
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='下颈线突破')

        """五笔三卖，要求bi5.low是最低点"""
        if min_low == bi5.low < bi5.high < min(bi1.low, bi3.low) \
                < max(bi1.low, bi3.low) < min(bi1.high, bi3.high) < max_high:
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='类三卖')

    return v


def check_seven_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别七笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的七笔
    :param di: 最近一笔为倒数第i笔
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='次级别无结构', v2='其他', v3='其他')

    if len(bis) != 7:
        return v

    bi1, bi2, bi3, bi4, bi5, bi6, bi7 = bis
    max_high = max([x.high for x in bis])
    min_low = min([x.low for x in bis])
    direction = bi7.direction

    assert direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if direction == Direction.Down:
        if bi1.high == max_high and bi7.low == min_low:
            # aAbcd式底背驰
            if min(bi2.high, bi4.high) > max(bi2.low, bi4.low) > bi6.high and bi7.power < bi5.power:
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

            # abcAd式底背驰
            if bi2.low > min(bi4.high, bi6.high) > max(bi4.low, bi6.low) and bi7.power < (bi1.high - bi3.low):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

            # aAb式底背驰
            if min(bi2.high, bi4.high, bi6.high) > max(bi2.low, bi4.low, bi6.low) and bi7.power < bi1.power:
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

            # 类趋势底背驰
            if bi2.low > bi4.high and bi4.low > bi6.high and bi7.power < max(bi5.power, bi3.power, bi1.power):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='底背驰')

        # 向上中枢完成
        if bi4.low == min_low and min(bi1.high, bi3.high) > max(bi1.low, bi3.low) \
                and min(bi5.high, bi7.high) > max(bi5.low, bi7.low) \
                and max(bi4.high, bi6.high) > min(bi3.high, bi4.high):
            if max(bi1.low, bi3.low) < max(bi5.high, bi7.high):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='向上中枢完成')

        # 七笔三买：1~3构成中枢，最低点在1~3，最高点在5~7，5~7的最低点大于1~3的最高点
        if min(bi1.low, bi3.low) == min_low and max(bi5.high, bi7.high) == max_high \
                and min(bi5.low, bi7.low) > max(bi1.high, bi3.high) \
                and min(bi1.high, bi3.high) > max(bi1.low, bi3.low):
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='类三买')

    if direction == Direction.Up:
        """# 顶背驰"""
        if bi1.low == min_low and bi7.high == max_high:
            """# aAbcd式顶背驰"""
            if bi6.low > min(bi2.high, bi4.high) > max(bi2.low, bi4.low) and bi7.power < bi5.power:
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

            """# abcAd式顶背驰"""
            if min(bi4.high, bi6.high) > max(bi4.low, bi6.low) > bi2.high and bi7.power < (bi3.high - bi1.low):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

            """# aAb式顶背驰"""
            if min(bi2.high, bi4.high, bi6.high) > max(bi2.low, bi4.low, bi6.low) and bi7.power < bi1.power:
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

            """# 类趋势顶背驰"""
            if bi2.high < bi4.low and bi4.high < bi6.low and bi7.power < max(bi5.power, bi3.power, bi1.power):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='顶背驰')

        """# 向下中枢完成"""
        if bi4.high == max_high and min(bi1.high, bi3.high) > max(bi1.low, bi3.low) \
                and min(bi5.high, bi7.high) > max(bi5.low, bi7.low) \
                and min(bi4.low, bi6.low) < max(bi3.low, bi4.low):
            if min(bi1.high, bi3.high) > min(bi5.low, bi7.low):
                return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='向下中枢完成')

        """# 七笔三卖：1~3构成中枢，最高点在1~3，最低点在5~7，5~7的最高点小于1~3的最低点"""
        if min(bi5.low, bi7.low) == min_low and max(bi1.high, bi3.high) == max_high \
                and max(bi7.high, bi5.high) < min(bi1.low, bi3.low) \
                and min(bi1.high, bi3.high) > max(bi1.low, bi3.low):
            return Signal(k1=freq.value, k2=di_name, k3='基础形态', v1='类三卖')
    return v


def check_nine_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别九笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的九笔
    :param di: 最近一笔为倒数第i笔
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='次级别无结构', v2='其他', v3='其他')
    if len(bis) != 9:
        return v

    direction = bis[-1].direction
    bi1, bi2, bi3, bi4, bi5, bi6, bi7, bi8, bi9 = bis
    max_high = max([x.high for x in bis])
    min_low = min([x.low for x in bis])
    assert direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if direction == Direction.Down:
        if min_low == bi9.low and max_high == bi1.high:
            # aAb式类一买
            if min(bi2.high, bi4.high, bi6.high, bi8.high) > max(bi2.low, bi4.low, bi6.low, bi8.low) \
                    and bi9.power < bi1.power and bi3.low >= bi1.low and bi7.high <= bi9.high:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # aAbcd式类一买
            if min(bi2.high, bi4.high, bi6.high) > max(bi2.low, bi4.low, bi6.low) > bi8.high \
                    and bi9.power < bi7.power:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # ABC式类一买
            if bi3.low < bi1.low and bi7.high > bi9.high \
                    and min(bi4.high, bi6.high) > max(bi4.low, bi6.low) \
                    and (bi1.high - bi3.low) > (bi7.high - bi9.low):
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # 类趋势一买
            if bi8.high < bi6.low < bi6.high < bi4.low < bi4.high < bi2.low \
                    and bi9.power < max([bi1.power, bi3.power, bi5.power, bi7.power]):
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

        # 九笔类一买（2~4构成中枢A，6~8构成中枢B，9背驰）
        if max_high == max(bi1.high, bi3.high) and min_low == bi9.low \
                and min(bi2.high, bi4.high) > max(bi2.low, bi4.low) \
                and min(bi2.low, bi4.low) > max(bi6.high, bi8.high) \
                and min(bi6.high, bi8.high) > max(bi6.low, bi8.low) \
                and bi9.power < bi5.power:
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

        # 类三买（1357构成中枢，最低点在3或5）
        if max_high == bi9.high > bi9.low \
                > max([x.high for x in [bi1, bi3, bi5, bi7]]) \
                > min([x.high for x in [bi1, bi3, bi5, bi7]]) \
                > max([x.low for x in [bi1, bi3, bi5, bi7]]) \
                > min([x.low for x in [bi3, bi5]]) == min_low:
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三买')

        # 类三买（357构成中枢，8的力度小于2，9回调不跌破GG构成三买）
        if bi8.power < bi2.power and max_high == bi9.high > bi9.low \
                > max([x.high for x in [bi3, bi5, bi7]]) \
                > min([x.high for x in [bi3, bi5, bi7]]) \
                > max([x.low for x in [bi3, bi5, bi7]]) > bi1.low == min_low:
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三买')

        if min_low == bi5.low and max_high == bi1.high and bi4.high < bi2.low:  # 前五笔构成向下类趋势
            zd = max([x.low for x in [bi5, bi7]])
            zg = min([x.high for x in [bi5, bi7]])
            gg = max([x.high for x in [bi5, bi7]])
            if zg > zd and bi8.high > gg:  # 567构成中枢，且8的高点大于gg
                if bi9.low > zg:
                    return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三买')

                # 类二买
                if bi9.high > gg > zg > bi9.low > zd:
                    return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类二买')

    if direction == Direction.Up:
        if max_high == bi9.high and min_low == bi1.low:
            """# aAbBc式类一卖"""
            if bi6.low > min(bi2.high, bi4.high) > max(bi2.low, bi4.low) \
                    and min(bi6.high, bi8.high) > max(bi6.low, bi8.low) \
                    and max(bi2.high, bi4.high) < min(bi6.low, bi8.low) \
                    and bi9.power < bi5.power:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# aAb式类一卖"""
            if min(bi2.high, bi4.high, bi6.high, bi8.high) > max(bi2.low, bi4.low, bi6.low, bi8.low) \
                    and bi9.power < bi1.power and bi3.high <= bi1.high and bi7.low >= bi9.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# aAbcd式类一卖"""
            if bi8.low > min(bi2.high, bi4.high, bi6.high) > max(bi2.low, bi4.low, bi6.low) \
                    and bi9.power < bi7.power:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# ABC式类一卖"""
            if bi3.high > bi1.high and bi7.low < bi9.low \
                    and min(bi4.high, bi6.high) > max(bi4.low, bi6.low) \
                    and (bi3.high - bi1.low) > (bi9.high - bi7.low):
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# 类趋势一卖"""
            if bi8.low > bi6.high > bi6.low > bi4.high > bi4.low > bi2.high \
                    and bi9.power < max([bi1.power, bi3.power, bi5.power, bi7.power]):
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

        """# 九笔三卖"""
        if max_high == bi1.high and min_low == bi9.low \
                and bi9.high < max([x.low for x in [bi3, bi5, bi7]]) < min([x.high for x in [bi3, bi5, bi7]]):
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三卖')

        if min_low == bi1.low and max_high == bi5.high and bi2.high < bi4.low:  # 前五笔构成向上类趋势
            zd = max([x.low for x in [bi5, bi7]])
            zg = min([x.high for x in [bi5, bi7]])
            dd = min([x.low for x in [bi5, bi7]])
            if zg > zd and bi8.low < dd:  # 567构成中枢，且8的低点小于dd
                if bi9.high < zd:
                    return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三卖')

                """# 类二卖"""
                if dd < zd <= bi9.high < zg:
                    return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类二卖')
    return v


def check_eleven_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别十一笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的十一笔
    :param di: 最近一笔为倒数第i笔
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='次级别无结构', v2='其他', v3='其他')
    if len(bis) != 11:
        return v

    direction = bis[-1].direction
    bi1, bi2, bi3, bi4, bi5, bi6, bi7, bi8, bi9, bi10, bi11 = bis
    max_high = max([x.high for x in bis])
    min_low = min([x.low for x in bis])
    assert direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if direction == Direction.Down:
        if min_low == bi11.low and max_high == bi1.high:
            # ABC式类一买，A5B3C3
            if bi5.low == min([x.low for x in [bi1, bi3, bi5]]) \
                    and bi9.low > bi11.low and bi9.high > bi11.high \
                    and bi8.high > bi6.low and bi1.high - bi5.low > bi9.high - bi11.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # ABC式类一买，A3B3C5
            if bi1.high > bi3.high and bi1.low > bi3.low \
                    and bi7.high == max([x.high for x in [bi7, bi9, bi11]]) \
                    and bi6.high > bi4.low and bi1.high - bi3.low > bi7.high - bi11.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # ABC式类一买，A3B5C3
            if bi1.low > bi3.low and min(bi4.high, bi6.high, bi8.high) > max(bi4.low, bi6.low, bi8.low) \
                    and bi9.high > bi11.high and bi1.high - bi3.low > bi9.high - bi11.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # a1Ab式类一买，a1（1~7构成的类趋势）
            if bi2.low > bi4.high > bi4.low > bi6.high > bi5.low > bi7.low and bi10.high > bi8.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

        # 类二买（1~7构成盘整背驰，246构成下跌中枢，9/11构成上涨中枢，且上涨中枢GG大于下跌中枢ZG）
        if bi7.power < bi1.power and min_low == bi7.low < max([x.low for x in [bi2, bi4, bi6]]) \
                < min([x.high for x in [bi2, bi4, bi6]]) < max([x.high for x in [bi9, bi11]]) < bi1.high == max_high \
                and bi11.low > min([x.low for x in [bi2, bi4, bi6]]) \
                and min([x.high for x in [bi9, bi11]]) > max([x.low for x in [bi9, bi11]]):
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类二买')

        # 类二买（1~7为区间极值，9~11构成上涨中枢，上涨中枢GG大于4~6的最大值，上涨中枢DD大于4~6的最小值）
        if max_high == bi1.high and min_low == bi7.low \
                and min(bi9.high, bi11.high) > max(bi9.low, bi11.low) \
                and max(bi11.high, bi9.high) > max(bi4.high, bi6.high) \
                and min(bi9.low, bi11.low) > min(bi4.low, bi6.low):
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类二买')

        # 类三买（1~9构成大级别中枢，10离开，11回调不跌破GG）
        gg = max([x.high for x in [bi1, bi2, bi3]])
        zg = min([x.high for x in [bi1, bi2, bi3]])
        zd = max([x.low for x in [bi1, bi2, bi3]])
        dd = min([x.low for x in [bi1, bi2, bi3]])
        if max_high == bi11.high and bi11.low > zg > zd \
                and gg > bi5.low and gg > bi7.low and gg > bi9.low \
                and dd < bi5.high and dd < bi7.high and dd < bi9.high:
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类三买')

    if direction == Direction.Up:
        if max_high == bi11.high and min_low == bi1.low:
            # ABC式类一卖，A5B3C3
            if bi5.high == max([bi1.high, bi3.high, bi5.high]) and bi9.low < bi11.low and bi9.high < bi11.high \
                    and bi8.low < bi6.high and bi11.high - bi9.low < bi5.high - bi1.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            # ABC式类一卖，A3B3C5
            if bi7.low == min([bi11.low, bi9.low, bi7.low]) and bi1.high < bi3.high and bi1.low < bi3.low \
                    and bi6.low < bi4.high and bi11.high - bi7.low < bi3.high - bi1.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            # ABC式类一卖，A3B5C3
            if bi1.high < bi3.high and min(bi4.high, bi6.high, bi8.high) > max(bi4.low, bi6.low, bi8.low) \
                    and bi9.low < bi11.low and bi3.high - bi1.low > bi11.high - bi9.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

        # 类二卖：1~9构成类趋势,11不创新高
        if max_high == bi9.high > bi8.low > bi6.high > bi6.low > bi4.high > bi4.low > bi2.high > bi1.low == min_low \
                and bi11.high < bi9.high:
            return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类二卖')
    return v


def check_thirteen_bi(bis: List[Union[BI, FakeBI]], freq: Freq, di: int = 1) -> Signal:
    """识别十三笔形态
    :param freq: K线周期，也可以称为级别
    :param bis: 由远及近的十三笔
    :param di: 最近一笔为倒数第i笔
    """
    di_name = f"倒{di}笔"
    v = Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='次级别无结构', v2='其他', v3='其他')
    if len(bis) != 13:
        return v

    direction = bis[-1].direction
    bi1, bi2, bi3, bi4, bi5, bi6, bi7, bi8, bi9, bi10, bi11, bi12, bi13 = bis
    max_high = max([x.high for x in bis])
    min_low = min([x.low for x in bis])

    assert direction in [Direction.Down, Direction.Up], "direction 的取值错误"

    if direction == Direction.Down:
        if min_low == bi13.low and max_high == bi1.high:
            # ABC式类一买，A5B3C5
            if bi5.low < min(bi1.low, bi3.low) and bi9.high > max(bi11.high, bi13.high) \
                    and bi8.high > bi6.low and bi1.high - bi5.low > bi9.high - bi13.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # ABC式类一买，A3B5C5
            if bi3.low < min(bi1.low, bi5.low) and bi9.high > max(bi11.high, bi13.high) \
                    and min(bi4.high, bi6.high, bi8.high) > max(bi4.low, bi6.low, bi8.low) \
                    and bi1.high - bi3.low > bi9.high - bi13.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

            # ABC式类一买，A5B5C3
            if bi5.low < min(bi1.low, bi3.low) and bi11.high > max(bi9.high, bi13.high) \
                    and min(bi6.high, bi8.high, bi10.high) > max(bi6.low, bi8.low, bi10.low) \
                    and bi1.high - bi5.low > bi11.high - bi13.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一买')

    if direction == Direction.Up:
        if max_high == bi13.high and min_low == bi1.low:
            """# ABC式类一卖，A5B3C5"""
            if bi5.high > max(bi3.high, bi1.high) and bi9.low < min(bi11.low, bi13.low) \
                    and bi8.low < bi6.high and bi5.high - bi1.low > bi13.high - bi9.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# ABC式类一卖，A3B5C5"""
            if bi3.high > max(bi5.high, bi1.high) and bi9.low < min(bi11.low, bi13.low) \
                    and min(bi4.high, bi6.high, bi8.high) > max(bi4.low, bi6.low, bi8.low) \
                    and bi3.high - bi1.low > bi13.high - bi9.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')

            """# ABC式类一卖，A5B5C3"""
            if bi5.high > max(bi3.high, bi1.high) and bi11.low < min(bi9.low, bi13.low) \
                    and min(bi6.high, bi8.high, bi10.high) > max(bi6.low, bi8.low, bi10.low) \
                    and bi5.high - bi1.low > bi13.high - bi11.low:
                return Signal(k1=freq.value, k2=di_name, k3='类买卖点', v1='类一卖')
    return v


# 以上是信号计算的辅助函数，主要是形态识别等。
# ----------------------------------------------------------------------------------------------------------------------
# 以下是信号计算函数（前缀固定为 get_s）

def get_s_three_bi(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """倒数第i笔的三笔形态信号
    :param c: CZSC 对象
    :param di: 最近一笔为倒数第i笔
    :return: 信号字典
    """
    assert di >= 1
    bis = c.finished_bis
    freq: Freq = c.freq
    s = OrderedDict()
    v = Signal(k1=str(freq.value), k2=f"倒{di}笔", k3="三笔形态", v1="次级别无结构", v2='其他', v3='其他')
    s[v.key] = v.value

    if not bis:
        return s

        
    if di == 1:
        three_bi = bis[-3:]
    else:
        three_bi = bis[-3 - di + 1: -di + 1]

    v = check_three_bi(three_bi, freq, di)
    s[v.key] = v.value
    return s


def get_s_base_xt(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """倒数第i笔的基础形态信号
    :param c: CZSC 对象
    :param di: 最近一笔为倒数第i笔
    :return: 信号字典
    """
    assert di >= 1

    bis = c.finished_bis
    freq: Freq = c.freq
    s = OrderedDict()
    v = Signal(k1=str(freq.value), k2=f"倒{di}笔", k3="基础形态", v1="次级别无结构", v2='其他', v3='其他')
    s[v.key] = v.value

    if not bis:
        return s

    if di == 1:
        five_bi = bis[-5:]
        seven_bi = bis[-7:]
    else:
        five_bi = bis[-5 - di + 1: -di + 1]
        seven_bi = bis[-7 - di + 1: -di + 1]

    for v in [check_five_bi(five_bi, freq, di), check_seven_bi(seven_bi, freq, di)]:
        if "次级别无结构" not in v.value:
            s[v.key] = v.value
    return s


def get_s_like_bs(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """倒数第i笔的类买卖点信号
    :param c: CZSC 对象
    :param di: 最近一笔为倒数第i笔
    :return: 信号字典
    """
    assert di >= 1
    bis = c.finished_bis
    freq: Freq = c.freq
    s = OrderedDict()
    v = Signal(k1=str(freq.value), k2=f"倒{di}笔", k3="类买卖点", v1="次级别无结构", v2='其他', v3='其他')
    s[v.key] = v.value

    if not bis:
        return s

    if di == 1:
        nine_bi = bis[-9:]
        eleven_bi = bis[-11:]
        thirteen_bi = bis[-13:]
    else:
        nine_bi = bis[-9 - di + 1: -di + 1]
        eleven_bi = bis[-11 - di + 1: -di + 1]
        thirteen_bi = bis[-13 - di + 1: -di + 1]

    for v in [check_nine_bi(nine_bi, freq, di), check_eleven_bi(eleven_bi, freq, di),
              check_thirteen_bi(thirteen_bi, freq, di)]:
        if "次级别无结构" not in v.value:
            s[v.key] = v.value
    return s


def get_s_bi_status(c: analyze.CZSC) -> OrderedDict:
    """倒数第1笔的表里关系信号
    :param c: CZSC 对象
    :return: 信号字典
    """
    freq: Freq = c.freq
    s = OrderedDict()
    v = Signal(k1=str(freq.value), k2="倒1笔", k3="表里关系", v1="其他", v2='其他', v3='其他')
    s[v.key] = v.value

    if c.bi_list:
        # 表里关系的定义参考：http://blog.sina.com.cn/s/blog_486e105c01007wc1.html
        min_ubi = min([x.low for x in c.bars_ubi])
        max_ubi = max([x.high for x in c.bars_ubi])

        last_bi = c.bi_list[-1]
        v = None
        if last_bi.direction == Direction.Down:
            if min_ubi < last_bi.low:
                v = Signal(k1=str(freq.value), k2="倒1笔", k3="表里关系", v1="向下延伸")
            else:
                v = Signal(k1=str(freq.value), k2="倒1笔", k3="表里关系", v1="底分完成")
        if last_bi.direction == Direction.Up:
            if max_ubi > last_bi.high:
                v = Signal(k1=str(freq.value), k2="倒1笔", k3="表里关系", v1="向上延伸")
            else:
                v = Signal(k1=str(freq.value), k2="倒1笔", k3="表里关系", v1="顶分完成")

        if v and "其他" not in v.value:
            s[v.key] = v.value
    return s


def get_s_d0_bi(c: analyze.CZSC) -> OrderedDict:
    """倒数第0笔信号
    :param c: CZSC 对象
    :return: 信号字典
    """
    freq: Freq = c.freq
    s = OrderedDict()

    default_signals = [
        Signal(k1=str(freq.value), k2="倒0笔", k3="方向", v1="其他", v2='其他', v3='其他'),
        Signal(k1=str(freq.value), k2="倒0笔", k3="长度", v1="其他", v2='其他', v3='其他'),
    ]
    for signal in default_signals:
        s[signal.key] = signal.value

    bis = c.finished_bis

    if bis:
        # 倒0笔方向
        last_bi = bis[-1]
        if last_bi.direction == Direction.Down:
            v = Signal(k1=str(freq.value), k2="倒0笔", k3="方向", v1="向上")
        elif last_bi.direction == Direction.Up:
            v = Signal(k1=str(freq.value), k2="倒0笔", k3="方向", v1="向下")
        else:
            raise ValueError

        if v and "其他" not in v.value:
            s[v.key] = v.value

        # 倒0笔长度
        bars_ubi = [x for x in c.bars_raw[-20:] if x.dt >= bis[-1].fx_b.elements[0].dt]
        if len(bars_ubi) >= 9:
            v = Signal(k1=str(freq.value), k2="倒0笔", k3="长度", v1="9根K线以上")
        elif 9 > len(bars_ubi) > 5:
            v = Signal(k1=str(freq.value), k2="倒0笔", k3="长度", v1="5到9根K线")
        else:
            v = Signal(k1=str(freq.value), k2="倒0笔", k3="长度", v1="5根K线以下")

        if "其他" not in v.value:
            s[v.key] = v.value
    return s


def get_s_di_bi(c: analyze.CZSC, di: int = 1) -> OrderedDict:
    """倒数第i笔的表里关系信号
    :param c: CZSC 对象
    :param di: 最近一笔为倒数第i笔
    :return: 信号字典
    """
    assert di >= 1
    freq: Freq = c.freq
    s = OrderedDict()
    k1 = str(freq.value)
    k2 = f"倒{di}笔"

    default_signals = [
        Signal(k1=k1, k2=k2, k3="方向", v1="其他", v2='其他', v3='其他'),
        Signal(k1=k1, k2=k2, k3="长度", v1="其他", v2='其他', v3='其他'),
        Signal(k1=k1, k2=k2, k3="拟合优度", v1="其他", v2='其他', v3='其他'),
    ]
    for signal in default_signals:
        s[signal.key] = signal.value

    bis = c.finished_bis
    if not bis:
        return s

    last_bi = bis[-di]

    # 方向
    v1 = Signal(k1=k1, k2=k2, k3="方向", v1=last_bi.direction.value)
    s[v1.key] = v1.value

    # 长度
    if len(last_bi.bars) >= 15:
        v = Signal(k1=k1, k2=k2, k3="长度", v1="15根K线以上")
    elif 15 > len(c.bars_ubi) > 9:
        v = Signal(k1=k1, k2=k2, k3="长度", v1="9到15根K线")
    else:
        v = Signal(k1=k1, k2=k2, k3="长度", v1="9根K线以下")

    if "其他" not in v.value:
        s[v.key] = v.value

    # 拟合优度
    rsq = RSQ([x.close for x in last_bi.bars[1:-1]])
    if rsq > 0.8:
        v = Signal(k1=k1, k2=k2, k3="拟合优度", v1="大于0.8")
    elif rsq < 0.2:
        v = Signal(k1=k1, k2=k2, k3="拟合优度", v1="小于0.2")
    else:
        v = Signal(k1=k1, k2=k2, k3="拟合优度", v1="0.2到0.8之间")

    if "其他" not in v.value:
        s[v.key] = v.value
    return 
    
def get_s_sma(c: analyze.CZSC, di: int = 1, t_seq=(27, 60, 108, 216)) -> OrderedDict:
    """获取倒数第i根K线的均线金叉情况"""
    freq: Freq = c.freq
    s = OrderedDict()

    k1 = str(freq.value)
    k2 = f"倒{di}K"
    x1 = Signal(k1=k1, k2=k2, k3=f"均线多空", v1="其他", v2='其他', v3='其他')
    s[x1.key] = x1.value

    n = max(t_seq) + 100
    if len(c.bars_raw) < n:
        return s
    
    bars = c.bars_nraw
    c_len = len(bars)
    ma27 = SMA(np.array([x.close for x in bars]), 27)
    ma60 = SMA(np.array([x.close for x in bars]), 60)
    ma108 = SMA(np.array([x.close for x in bars]), 108)
    ma216 = SMA(np.array([x.close for x in bars]), 216)
    ma2000 = SMA(np.array([x.close for x in bars]), 2000)

    close = bars[- 1].close
    line27 = ma27[-1]
    line60 = ma60[-1]
    line108 = ma108[-1]
    line216 = ma216[-1]
    line2000 = ma2000[-1]
    
    cross_point1 = find_cross_point(ma27, ma60, c_len)
    cross_point2 = find_cross_point(ma27, ma108, c_len)
    
    cross_point4 = find_cross_point(ma60, ma108, c_len)

    min_cross = cross_point4 > cross_point2 and cross_point2 > cross_point1
    
    lineup = line27 > line60 and line27 > line108 and line27 > line216#3大于7，大于13 大于 27日均线
    
    priceup = close > line108 and close > line216#收盘价过13日和250日均线
    
    cross_flag = not (cross_point1 == c_len) and min_cross 
    
    if lineup and priceup and cross_flag:
       v1 = Signal(k1=k1, k2=k2, k3=f"均线多空", v1="多头")
    else:
        v1 = Signal(k1=k1, k2=k2, k3=f"均线多空", v1="空头")
    s[v1.key] = v1.value
    return s
            
        
    
    