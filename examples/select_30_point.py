#! /usr/bin/env python
# -*- coding: utf-8 -*-

import baostock as bs
import pandas as pd
import pymysql
import time
import datetime
import calendar
from datetime import timedelta
from itertools import chain
import os
import pandas as pd
from collections import OrderedDict
from czsc import CZSC, CzscAdvancedTrader, Freq
from czsc.utils import BarGenerator
from czsc import signals
from czsc.objects import RawBar
from czsc.enum import Freq
from czsc import signals
from czsc.signals.signals import get_default_signals, get_s_three_bi, get_s_d0_bi
from czsc.signals.ta import get_s_single_k, get_s_three_k, get_s_sma, get_s_macd
from czsc.signals.bxt import get_s_like_bs, get_s_d0_bi, get_s_bi_status, get_s_di_bi, get_s_base_xt, get_s_three_bi

import new_signal
import datetime
import openpyxl
from openpyxl import load_workbook
    
import os,re,time
import pathlib
from czsc.utils.cache import home_path

os.environ['czsc_verbose'] = "1"        # 是否输出详细执行信息，0 不输出，1 输出
os.environ['czsc_min_bi_len'] = "6"     # 通过环境变量设定最小笔长度，6 对应新笔定义，7 对应老笔定义
pd.set_option('mode.chained_assignment', None)
pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', 20)

freq_map = {'1min': Freq.F1, '5min': Freq.F5, '15min': Freq.F15, '30min': Freq.F30,
            '60min': Freq.F60, 'D': Freq.D, 'W': Freq.W, 'M': Freq.M}


def float2(value, num):
    #print(round(float(value),num))
    return round(float(value),num)
    
    
def download_30_data_by_code(code, start_date, end_date):
    # 获取指定股票数据
    data_df = pd.DataFrame()
    #print(stockcodes.shape[0])
    if True:
    # for code in stockcodes["code"]:
        print("Downloading day:" + code)
        # k_rs = bs.query_history_k_data_plus("sz.003037", "date,time,code,open,high,low,close,volume,amount,adjustflag", start_date, end_date, frequency="30", adjustflag="3")
        k_rs = bs.query_history_k_data_plus(code, "date,time,code,open,high,low,close,volume,amount,adjustflag", start_date, end_date, frequency="30", adjustflag="2")
        ##sh.600006
        # k_rs = bs.query_history_k_data_plus("sh.600006", "date,code,open,high,low,close,preclose,volume,amount,adjustflag,turn,tradestatus,pctChg,isST", start_date, end_date, frequency="d", adjustflag="2")
        df_stockload = k_rs.get_data()
        c_len = df_stockload.shape[0]
        freq = '30min'
        bars = []
        if c_len > 0:
            #print(df_stockload)
            # print(df_stockload['close'][c_len - 1])
            for j in range(c_len):
                #print(df_stockload['amount'][j])
                try:
                    str = df_stockload['time'][j]
                    dt = datetime.datetime.strptime(str, "%Y%m%d%H%M%S%f")
                    bars.append(RawBar(symbol=df_stockload['code'][j], dt=dt, id=j, freq=freq_map[freq],
                                   open=round(float(df_stockload['open'][j]), 2),
                                   close=round(float(df_stockload['close'][j]), 2),
                                   high=round(float(df_stockload['high'][j]), 2),
                                   low=round(float(df_stockload['low'][j]), 2),
                                   vol=round(float(df_stockload['volume'][j]), 2)))
                except Exception as err:
                     print(err)
                     bars = []
                    #print(str(resu[0]),str(resu[1]))
                     return bars
            #print(c_len)
            return bars
        else:
            return bars
            
def to_csv(export_signals):
    """保存到表格中"""

    # 新建文件的名称
    new_excel_file = home_path + '/30_struct_stock_signals.xlsx'

    # 新建sheet的名称
    ws1_name = 'stock_signals'
    # ws2_name = 'testsheet'

    # print(new_excel_file)
    # 创建一个excel表格，默认表格内存在一个名为‘Sheet’的sheet，
    # 表格创建在缓存区，通过save(‘file_name.xlsx’)函数保存到指定目录下
    if not pathlib.Path(new_excel_file).exists():
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(new_excel_file)
    # 将默认创建的名为‘Sheet’的sheet重命名为ws1_name
    if 'Sheet' in wb.sheetnames:
        wb['Sheet'].title = ws1_name
    
    sheet1 = wb.active
    #print(sheet1)
    
    row_file = 5 # 生成5行
    col_three = 3 # 生成3列

    for row in range(len(export_signals)):
        for col in range(3):
            rw = row + 1
            cl = col + 1
            sheet1.cell(row=rw, column=cl, value=export_signals[row][col])

    # create new sheet,sheet name = ws2_name
    # wb.create_sheet(ws2_name)

    ws=wb[ws1_name]
               
    # print(new_excel_file)
    wb.save(new_excel_file)#保存 
    wb.close()

    
def get_all_stock(date):
    stock_rs = bs.query_all_stock(date)
    stock_df = stock_rs.get_data()
    #print(stock_df)
    return stock_df
    

                    
def get_last_trade_day(start_dt, end_dt):
    rs = bs.query_trade_dates(start_date=start_dt, end_date=end_dt)
    print('query_trade_dates respond error_code:'+rs.error_code)
    print('query_trade_dates respond  error_msg:'+rs.error_msg)

    #### 打印结果集 ####
    data_list = []


    while (rs.error_code == '0') & rs.next():
        # 获取一条记录，将记录合并在一起
        data = rs.get_row_data()
        #print(data[1])
        if data[1] == '1':
            data_list.append(data)
        
    #result = pd.DataFrame(data_list, columns=rs.fields)

    #### 结果集输出到csv文件 ####   
    #result.to_csv("D:\\trade_datas.csv", encoding="gbk", index=False)
    #print(data_list)
    print(data_list[len(data_list) - 1][0])
    return data_list[len(data_list) - 1][0]
    
def get_test_signals(c: CZSC) -> OrderedDict:
        s = OrderedDict({"symbol": c.symbol, "dt": c.bars_raw[-1].dt, "close": c.bars_raw[-1].close})
        s.update(get_s_d0_bi(c))
        s.update(get_s_three_k(c, 1))
        s.update(get_s_di_bi(c, 1))
        s.update(get_s_macd(c, 1))
        return s
        


# 定义一些需要观察的信号，可以是多级别同时计算
def get_simple_signals(cat: CzscAdvancedTrader) -> OrderedDict:
    s = OrderedDict({"symbol": cat.symbol, "dt": cat.end_dt, "close": cat.latest_price})
    for _, c in cat.kas.items():
        if c.freq == Freq.F30:
            s.update(new_signal.get_s_base_xt(c, di=1))
            s.update(new_signal.get_s_three_bi(c, di=1))
            s.update(new_signal.get_s_like_bs(c, di=1))

    return s


def simple_strategy(symbol):
    return {"symbol": symbol, "get_signals": get_simple_signals}



    
if __name__ == '__main__':
    while True:
        
        time_temp = datetime.datetime.now() - datetime.timedelta(days=300)
        start_dt = time_temp.strftime('%Y-%m-%d')
        
        time_temp = datetime.datetime.now() - datetime.timedelta(days=1)
        end_dt = time_temp.strftime('%Y-%m-%d')
        
        
        bs.login()
        
        lasttradeday = get_last_trade_day(start_dt, end_dt)
        weekday = datetime.datetime.now().isoweekday()
        
        
        # 获取全部股票的信息
        all_stock_codes = get_all_stock(lasttradeday)
        # print(all_stock_codes)
        
        total = all_stock_codes.shape[0]
        # 获取全部股票的日信息
        if total != 0:
            print("获取全部股票的日信息")
            export_signals = []           
            for code in all_stock_codes["code"]:
                bars = []
                if ("sh.0" not in code) and ("bj.8" not in code)  and ("bj.4" not in code) and ("sh.688" not in code) and ("sz.399" not in code):
                    bars = download_30_data_by_code(code, start_dt, end_dt)
                    if bars == []:
                        continue
                    else:
                        c = CZSC(bars, get_signals=get_test_signals)
                        
                        #c.open_in_browser()
                        # K线合成器，这是多级别联立分析的数据支撑。示例为从日线逐K合成周线、月线
                        bg = BarGenerator(base_freq='30分钟', freqs=[], max_count=5000)
                        for bar in bars:
                            bg.update(bar)
                       
                        cat = CzscAdvancedTrader(bg, simple_strategy)
                        signals = {k: v for k, v in cat.s.items() if len(k.split("_")) == 3}
                        for k in signals.keys():
                            #export_signals.append([cat.symbol, k, signals[k]])
                            if ("次级别无结构" not in signals[k]) :
                                print(cat.symbol + " " + k + " " + signals[k])
                                export_signals.append([cat.symbol, k, signals[k]])
                    
                    
                
                #cat.open_in_browser()
            to_csv(export_signals)
                
            
        print("handle finish")

        
